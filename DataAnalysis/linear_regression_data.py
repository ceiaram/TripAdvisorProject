import pandas as pd
import openpyxl
from openpyxl.utils.exceptions import IllegalCharacterError
from lxml import etree
from collections import Counter
import itertools 
import re
import numpy as np
import nltk

# # download the punkt tokenizer if it's not already installed
# nltk.download('punkt')


# Create a new workbook
workbook = openpyxl.Workbook()
# Select the active worksheet
worksheet = workbook.active
def update_excel(data_dict, worksheet, skip_columns_after=None):
        # Clear the existing data in the worksheet
        worksheet.delete_rows(1, worksheet.max_row)

        # Write the headers to the first row
        headers = list(data_dict.keys())
        col_index = 1
        for header in headers:
            if header in skip_columns_after:
                worksheet.cell(row=1, column=col_index).value = header
                worksheet.cell(row=1, column=col_index + 1).value = ''
                worksheet.cell(row=1, column=col_index + 2).value = ''
                col_index += 3
            else:
                worksheet.cell(row=1, column=col_index).value = header
                col_index += 1

        # Pad the shorter lists with empty values
        max_length = max([len(data_dict[key]) for key in headers])
        for key, value in data_dict.items():
            if len(value) < max_length:
                data_dict[key] += [''] * (max_length - len(value))

        # Write the data to the remaining rows
        for i in range(max_length):
            col_index = 1
            for header in headers:
                if header in skip_columns_after:
                    worksheet.cell(row=i+2, column=col_index).value = data_dict[header][i]
                    worksheet.cell(row=i+2, column=col_index + 1).value = ''
                    worksheet.cell(row=i+2, column=col_index + 2).value = ''
                    col_index += 3
                else:
                    worksheet.cell(row=i+2, column=col_index).value = data_dict[header][i]
                    col_index += 1

        # Check if there are any invalid XML elements in the workbook
        with open("linear_regression_data.xlsx", "rb") as f:
            try:
                etree.parse(f)
            except etree.XMLSyntaxError as e:
                print(f"\nInvalid XML element found in workbook: {e.args[0]}")

        # Save the workbook
        workbook.save("linear_regression_data.xlsx")

if __name__ == "__main__":
    # number of words,number of sentences, convert manger response to boolean, convert customer rating to #
    df = pd.read_excel('data_analysis.xlsx', usecols=['Customer Rating', 'Reviews','Manager Response', 'Sustainabilty Score', 'Helpful Vote', 'Contributions'])

    result = {'Customer Rating' :[],'Number of Words in Review':[], 'Number of Sentences in Review' :[],'Manager Response':[], 'Sustainabilty Score':[],
              'Helpful Vote':[], 'Contributions':[]}
    
    helpful_votes = []
    manager_responses = []
    
    # loop through each row of the DataFrame and extract the rating value using regex
    for index, row in df.iterrows():
        # Extract the numeric value only from format '3.0 out of 5.0 bubbles'
        rating = re.findall(r'\d+\.\d+', row['Customer Rating'])[0]
        result['Customer Rating'].append(float(rating))

        # Get the number of words found in a customer review
        review = row['Reviews']
        num_words = len(review.split())
        result['Number of Words in Review'].append(int(num_words))

        # split the text into sentences using nltk
        sentences = nltk.sent_tokenize(review)
        result['Number of Sentences in Review'].append(len(sentences))
        
        # Handle cases where the Helpful Vote column contains the string "none"
        if row['Helpful Vote'] == 'none':
            helpful_votes.append(0)
        else:
            helpful_votes.append(int(re.findall(r'\d+', row['Helpful Vote'])[0]))
        
        # Convert the Manager Response column to boolean values
        if row['Manager Response'] == 'none':
            manager_responses.append(0)
        else:
            manager_responses.append(1)

        result['Sustainabilty Score'].append(row['Sustainabilty Score'])
        result['Contributions'].append(row['Contributions'])
        
    result['Manager Response'] = manager_responses
    result['Helpful Vote'] = helpful_votes
    
    # create a new DataFrame from the result dictionary
    result_df = pd.DataFrame(result)

    # display the result DataFrame
    print(result_df)
    update_excel(result, worksheet, skip_columns_after=['Contributions'])


