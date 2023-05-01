import pymongo
import dns # required for connecting with SRV
import pandas as pd
import spacy
import re
import csv
import openpyxl
from collections import Counter
# Import the Matcher
from spacy.matcher import PhraseMatcher
from spacy.matcher import Matcher
from openpyxl.utils.exceptions import IllegalCharacterError
from lxml import etree


# Connect to database, 24hr timeout
client = pymongo.MongoClient("mongodb+srv://ceiaram:Xn704K3WQ0So6s09@tripadvisordb.qw0eojr.mongodb.net/?retryWrites=true&w=majority", socketTimeoutMS=86400000)
db = client.test

# Load the Excel workbook
workbook = openpyxl.load_workbook('dummy.xlsx')
# Select the sheet where the columns are located
worksheet = workbook['Sheet']

# Get database named "OrangeCountyHotels"
db = client["OrangeCountyHotels"]
hotel_info = {}
environmental_keywords = {
    'automatic light': 1,  # TWO-WORDED KEYWORDS MUST COME BEFORE ONE-WORDED KEY WORDS THAT ARE FOUND IN THE TWO-WORDED KEYWORDS OTHERWIS IT WON'T PICK UP THE TWO-WORDED KEYWORDS/CCORRECT SCORES 
    'clean air': 1,
    'clean environment': 1,
    'energy sav': 1,
    'environmental certificat': 1,
    'environmental concern': 1,
    'environmental manage': 1,
    'ev charg': 1,
    'ev park': 1,
    'electronic car': 1,
    'electric car': 1,
    'iso 14001': 1,
    'towel reuse': 1,
    'led light': 1,
    'leed certificat': 1,
    'locally grown': 1,
    'linen reuse': 1,
    'motion sensor': 1,
    'green space': 1,
    'green globe': 1,
    'green hotel': 1,
    'green issue': 1,
    'green lodge': 1,
    'green practice': 1,
    'green certificat': 1,
    'recharge station': 1,
    'sustainable material': 1,
    'sustainable land': 1,
    'sustainable transport': 1,
    'waste conservat': 1,
    'water sav': 1,
    'vehicle charg': 1,
    'green': 1,
    'conservat': 1,
    'eco': 1,
    'environment': 1,
    'reuse': 1,
    'recycle': 1,
    'renewable': 1,
    'solar': 1,
    'organic': 1,
    'biodegrad': 1,
    'biodiversit': 1,
    'degrad': 1,
    'hybrid': 1,
    'nature': 1,
    'efficient': 1,
    'vegetarian': 1,
    'vegan': 1,
    'healthy': 1,
    'sustain': 1,  
    'leed': 1,
    'certificat': 1,
    'non-fuel': 1,
    'reduce': 1
    }
social_keywords = {
    'corporate responsibilit': 1,
    'community support': 1,
    'ethical practice': 1,
    'local employee': 1,
    'local purchase': 1,
    'local community': 1,
    'responsible consumer behavior': 1,
    'responsible customer behavior': 1,
    'social responsibilit': 1,
    'sustainable communit': 1,
    'communit': 1,
    'responsibl': 1,
    'involvement': 1,
    }

cultural_keywords = {
    'cultural sustainabilit': 1,
    'safe work': 1,
    'work safe' : 1,
    'heritage': 1,
    'inclusi': 1,
    'traditional': 1,
    'diversit': 1,
    'equit': 1,
    'fair': 1,
    'dei': 1,
    }

economic_keywords = {
    'market demand': 1,
    'competitive': 1,
    'advantage': 1,
    'innovat': 1,
    'profit': 1,
    'investment': 1, #Need to check 
    'value': 1,
    'renovat': 1
    }

policy_keywords = {
    'corporate social responsibilit': 1,
    'environmental regulation': 1,
    'incentive program': 1,#Need to check 
    'labor regulat': 1,
    'organizational identit': 1,
    'social responsibilit': 1,
    'social justice': 1,
    'tax polic': 1,
    'zoning polic': 1,
    'stakeholder': 1,
    'regulation': 1,
    'policy': 1, #Need to check 
    'policies': 1, #Need to check 
    'transparen': 1   
    }

about_section_scores = []
def findKeywords(text, keywords):
       # create a regular expression pattern to match the whole word that contains one of the keywords(ex: Disney should find Disneyland and Disney-land and landdisney)
       pattern2 = re.compile(r'\b(\w*{}[\w]*)\b'.format('|'.join(re.escape(keyword) for keyword in keywords)), re.IGNORECASE)

       # create a regular expression pattern matches any word that starts with one of the keywords and is followed by zero or more word characters
       # main difference with previous pattern is this pattern does not capture the matched text, re.IGNORECASE is for disregard case sensitivity 
       pattern3 = re.compile(r'\b(?:{})\w*\b'.format('|'.join(map(re.escape, keywords))), re.IGNORECASE)
        
       word_variations = []
       #store the two-worded keywords and their respective values
       two_worded_keywords = {}

       # Includes two-worded keywords in the dictionary and adds their value to the score if both words are found in the text
       for keyword in keywords:
        words = keyword.split()
        if len(words) > 1:
            first_word = words[0]
            second_word = words[1]
            word_variations.append(r'\b{}\w*\b.*?\b{}\w*\b|\b{}\w*\b \w*\b'.format(re.escape(keyword), re.escape(second_word), re.escape(first_word)))
            two_worded_keywords[keyword] = keywords[keyword]

       # Pattern matches either a single word from the dictionary keys, or a two-word phrase that has been split into its constituent words and combined with the \w*
       pattern = re.compile(r'\b(?:{})\w*\b|\b(?:{})\b'.format('|'.join(map(re.escape, keywords.keys())), '|'.join(word_variations)), re.IGNORECASE)

       # Scores for each category 
       score = 0

       # search for a keyword in the text from pattern
       match = pattern.findall(text)
    #    Don't need pattern 2 or 3
    #    match2 = pattern2.findall(text)
    #    match3= pattern3.findall(text)

       matchesFound = []  
       matchesFound.extend(match)
    #    matchesFound.extend(match2)
    #    matchesFound.extend(match3)

       # remove duplicates from the list of keyword matches
       matchesFound = list(set(matchesFound))

       printed_matches = set()

       # Loop over all matches found and add correct values based on keywords
       for match in matchesFound:
            # Check if the match is a one-worded keyword and add its value to the score
            for keyword, value in keywords.items():
                if keyword in match.lower() and match not in printed_matches:
                    score += value
                    print(f"Matched keyword '{keyword}' with value {value}: '{match}'")
                    printed_matches.add(match)
            # Check if the match is a two-worded keyword and add its value to the score 
            for keyword, value in two_worded_keywords.items():
                if keyword in match.lower() and match not in printed_matches:
                    score += value
                    print(f"Matched keyword '{keyword}' with value {value}: '{match}'")
                    printed_matches.add(match)

       # Append scores to correct list to be added 
       about_section_scores.append(score)
           
       if printed_matches:
        print("Keywords found: ", printed_matches)
        print("Score: ", score)
       else:
        print('No keywords found.')
        printed_matches.add('none')
        
       return printed_matches




# Get collection
collection_hotels = db['HotelData']
for x in range(587):
    my_query = {'Index' : x}
    hotel_doc = collection_hotels.find(my_query)
     # Print Hotel with query 
    for doc in hotel_doc:
                # Obtain the hotel fields for the each hotel 
                hotel_name = doc['Hotel Name']
                # hotel_names.append(hotel_name)
                if doc['Hotel About Section']:
                    about_section = doc['Hotel About Section']
                    # if the about section is a string, access its elements
                    if isinstance(about_section, str):
                                        #Find keywords in hotel about section from each category of dictionary 
                                        enviornmentalResultsHotelDes = findKeywords(about_section, environmental_keywords)
                                        socialResultsHotelDes = findKeywords(about_section, social_keywords)
                                        culturalResultsHotelDes = findKeywords(about_section, cultural_keywords)
                                        economicResultsHotelDes = findKeywords(about_section, economic_keywords)
                                        policyResultsHotelDes = findKeywords(about_section, policy_keywords)
                                        

                                        # Get the sum of all the scores found in each category for the about section 
                                        sum_scores_about = 0
                                        print("ABOUT SECTION SCORES", about_section_scores)
                                        for scores in about_section_scores:
                                            sum_scores_about+= scores
                    else:
                                    about_section = "none"

                    hotel_info[hotel_name] = sum_scores_about
                    print(hotel_name, ":", sum_scores_about)
                about_section_scores.clear()


# for name, scores in hotel_info.items():
#     print(f'{name}: {scores}')
# print(len(hotel_info))






for name, scores in hotel_info.items():
   if name != 'none':
        print("Finding data in row...for: ", name)
        # Find the row numbers of the rows with the specified name
        row_numbers = []
        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=1):
            if row[0].value == name:
                row_numbers.append(row[0].row)

        print("Getting data in column...")
        # If the rows are found, get the data in column D for each row
        if row_numbers:
            column_data = []
            for row_number in row_numbers:
                # Write the score to column D for the current row
                worksheet.cell(row=row_number, column=4).value = scores

            # # Print the column data
            # print(column_data)
            # print(len(column_data))
        else:
            print("Rows not found")
   else:
       print("Hotel name is none, so skip: ", name)




# Save the updated workbook
workbook.save('dummy.xlsx')

# # Update the values in the column
# for i in range(2, worksheet.max_row + 1):
#     worksheet['Hotel About Section Sustainabilty Score' + str(i)] = 'New Value'

