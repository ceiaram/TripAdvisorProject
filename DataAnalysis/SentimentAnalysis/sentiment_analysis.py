#  The most common approach to scoring sentiment is to assign a numerical value to each sentiment category, such as -1 for negative, 0 for neutral, and +1 for positive
import nltk
import pandas as pd
import openpyxl
from lxml import etree
from nltk.sentiment import SentimentIntensityAnalyzer

# Create a new workbook
workbook = openpyxl.Workbook()
# Select the active worksheet
worksheet = workbook.active
sia = SentimentIntensityAnalyzer()

def is_positive(review: str) -> bool:
    # True if review has positive compound sentiment, False otherwise
    return sia.polarity_scores(review)["compound"] > 0
# print(sia.polarity_scores(string))

# print(is_positive(string))

# The compound score is calculated differently. It’s not just an average, and it can range from -1 to 1.

def update_excel(data_dict, worksheet, skip_columns_after=None):
        # Clear the existing data in the worksheet
        worksheet.delete_rows(1, worksheet.max_row)

        # Write the headers to the first row
        headers = list(data_dict.keys())
        col_index = 1
        for header in headers:
            if header in skip_columns_after:
                worksheet.cell(row=1, column=col_index).value = header
                worksheet.cell(row=1, column=col_index + 1).value = ''
                worksheet.cell(row=1, column=col_index + 2).value = ''
                col_index += 3
            else:
                worksheet.cell(row=1, column=col_index).value = header
                col_index += 1

        # Pad the shorter lists with empty values
        max_length = max([len(data_dict[key]) for key in headers])
        for key, value in data_dict.items():
            if len(value) < max_length:
                data_dict[key] += [''] * (max_length - len(value))

        # Write the data to the remaining rows
        for i in range(max_length):
            col_index = 1
            for header in headers:
                if header in skip_columns_after:
                    worksheet.cell(row=i+2, column=col_index).value = data_dict[header][i]
                    worksheet.cell(row=i+2, column=col_index + 1).value = ''
                    worksheet.cell(row=i+2, column=col_index + 2).value = ''
                    col_index += 3
                else:
                    worksheet.cell(row=i+2, column=col_index).value = data_dict[header][i]
                    col_index += 1

        # Check if there are any invalid XML elements in the workbook
        with open("sentiment_analysis_data.xlsx", "rb") as f:
            try:
                etree.parse(f)
            except etree.XMLSyntaxError as e:
                print(f"\nInvalid XML element found in workbook: {e.args[0]}")

        # Save the workbook
        workbook.save("sentiment_analysis_data.xlsx")


if __name__ == "__main__":
    df = pd.read_excel('data_analysis.xlsx', usecols=['Reviews','Sustainabilty Score'])
    result = {'Sustainabilty Score':[], 'Sentiment Negative Score':[], 'Sentiment Positive Score':[],
            'Sentiment Neutral Score':[], 'Sentiment Negative + Neutral Score':[], 'Sentiment Positive + Neutral Score':[]}


    # loop through each row of the DataFrame and extract the rating value using regex
    for index, row in df.iterrows():
        sustainability_score = row['Sustainabilty Score']
        result['Sustainabilty Score'].append(int(sustainability_score))

        # Get the setiment score for each rreview
        review = row['Reviews']
        scores = sia.polarity_scores(review)
        # extract the negative sentiment score
        neg_score = scores['neg']
        # extract the neutral sentiment score
        neutral_score = scores['neu']
        # extract the positive sentiment score
        positive_score = scores['pos']

        result['Sentiment Negative Score'].append(float(neg_score))
        result['Sentiment Neutral Score'].append(float(neutral_score))
        result['Sentiment Positive Score'].append(float(positive_score))
        result['Sentiment Negative + Neutral Score'].append(float(neg_score + neutral_score))
        result['Sentiment Positive + Neutral Score'].append(float(positive_score + neutral_score))
        print(review)
        print("SCORES: ", scores)
        print('Negative Score: ', neg_score)
        print('Neutral Score: ', neutral_score)
        print('Positive Score: ', positive_score)


    # create a new DataFrame from the result dictionary
    result_df = pd.DataFrame(result)
    update_excel(result, worksheet, skip_columns_after=['Sentiment Positive + Neutral Score'])
