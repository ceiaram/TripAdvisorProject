import pandas as pd
import openpyxl
from openpyxl.utils.exceptions import IllegalCharacterError
from lxml import etree
from collections import Counter
from operator import itemgetter
import itertools 
import re
import numpy as np

# Create a new workbook
workbook = openpyxl.Workbook()
# Select the active worksheet
worksheet = workbook.active

# global var
# total_hotels_price_above_401 = 0
# total_hotels_price_301_400 = 0
price_range_count_global = {}


class Analysis:
    # Get the top 10 hotels based on provided column data
    def get_top_10(df, col):
        # Drop rows with 'none' in the 'Hotel Name' column
        df = df[df['Hotel Name'] != 'none']
        # Sort the DataFrame by the specified column in descending order and extract the top 10 rows
        top_10 = df.sort_values(col, ascending=False).head(10)
        return top_10
    
    def get_hotel_keywords(df, top_10):
        top_10_list = top_10['Hotel Name'].tolist()
        results = []

        # Loop over the top 10 hotels and get keywords
        for hotel in top_10_list:
            # Get the rows with specific hotel name
            specified_rows = df[df.iloc[:,0] == hotel]
            keywords_list = []

            for i in range(len(specified_rows)):
                # Get the keywords for the current row
                current_row_keywords = []
                if specified_rows.iloc[i]['Enviornmental Keywords'] != 'none':
                    current_row_keywords.append((specified_rows.iloc[i]['Enviornmental Keywords'], 'Environmental'))

                if specified_rows.iloc[i]['Environmental Keywords[Environment]'] != 'none':
                    current_row_keywords.append((specified_rows.iloc[i]['Environmental Keywords[Environment]'], '[Environment]'))
                if specified_rows.iloc[i]['Environmental Keywords[Certificate]'] != 'none':
                    current_row_keywords.append((specified_rows.iloc[i]['Environmental Keywords[Certificate]'], '[Certificate]'))
                if specified_rows.iloc[i]['Environmental Keywords[Green Practices]'] != 'none':
                    current_row_keywords.append((specified_rows.iloc[i]['Environmental Keywords[Green Practices]'], '[Green Practices]'))
                if specified_rows.iloc[i]['Environmental Keywords[Sustainable Transportation]'] != 'none':
                    current_row_keywords.append((specified_rows.iloc[i]['Environmental Keywords[Sustainable Transportation]'], '[Sustainable Transportation]'))

                if specified_rows.iloc[i]['Social Keywords'] != 'none':
                    current_row_keywords.append((specified_rows.iloc[i]['Social Keywords'], 'Social'))

                if specified_rows.iloc[i]['Cultural Keywords'] != 'none':
                    current_row_keywords.append((specified_rows.iloc[i]['Cultural Keywords'], 'Cultural'))

                if specified_rows.iloc[i]['Economic Keywords'] != 'none':
                    current_row_keywords.append((specified_rows.iloc[i]['Economic Keywords'], 'Economic'))

                if specified_rows.iloc[i]['Policy Keywords'] != 'none':
                    current_row_keywords.append((specified_rows.iloc[i]['Policy Keywords'], 'Policy'))

                # Append the keywords to the list of keywords for the hotel
                keywords_list.extend(current_row_keywords)

            # Add the list of keywords for the hotel to the results list
            results.append((hotel, keywords_list))

        return results

    def get_top_10_keywords(df, col):
        top_10_keywords = df[col].tolist()
        # Exclude 'none', so it won't be counted as a common word
        top_10_keywords = [word for word in top_10_keywords if word != "none"]
        word_counts = Counter(top_10_keywords)

        result = []
        for word, count in word_counts.most_common(10):
            result.append((word, count))
        return result
    
    # Get top 5 most sustainable hotels in each hotel star rating
    def get_top_5_stars(df):
        # Sort the DataFrame by both the Hotel Rating and Hotel Sustainability Average columns
        df_sorted = df.sort_values(['Hotel Rating', 'Hotel Sustainabilty Average'], ascending=[True, False])

        # Group the sorted DataFrame by Hotel Rating and extract the top 5 hotels from each group
        top_5 = df_sorted.groupby('Hotel Rating').head(5)

        # Create a dictionary to store the top 5 hotels for each rating
        results_dict = {}
        # Extract the top 5 hotels for each rating and add them to the dictionary
        for rating in df['Hotel Rating'].unique():
            hotels = top_5.loc[top_5['Hotel Rating'] == rating][['Hotel Name', 'Hotel Sustainabilty Average', 'Total Number of Reviews', 'Hotel About Section Keywords']]
            hotels_list = [(hotel[0], hotel[1], hotel[2], hotel[3]) for hotel in hotels.values]
            results_dict[rating] = hotels_list
        return results_dict

    def get_total_price_range(df):
        # Convert the 'Hotel Price Range' column to string type
        df['Hotel Price Range'] = df['Hotel Price Range'].astype(str)

        # Extract numeric values from the 'Hotel Price Range' column using regular expression
        price_range_regex = r'\$(\d+)\s*-\s*\$(\d+)'
        # Extracts numeric value from the 'Hotel Price Range' column using the str.extract() method and converts it to an integer using the astype() method
        df[['Price Range Low', 'Price Range High']] = df['Hotel Price Range'].str.extract(price_range_regex).astype(float)
        df.loc[df['Hotel Price Range'].str.lower().str.contains('none'), ['Price Range Low', 'Price Range High']] = np.nan
        df['Price Range Average'] = df[['Price Range Low', 'Price Range High']].mean(axis=1, skipna=True)
        
        # Create a new column that categorizes the hotel prices into the price ranges using 'pd.cut' function
        df['Price Range Category'] = pd.cut(df['Price Range Average'], 
                                            bins=[0, 100, 200, 300, 400, np.inf], 
                                            labels=['Below $100', '$100-$200', '$201-$300', '$301-$400', 'Above $400'])
        
       
        # Count the number of hotels in each price range category
        price_range_count = df.groupby('Price Range Category')['Hotel Name'].count().reset_index(name='Number of Hotels')
        return price_range_count


    def get_top_5_by_price_range(df):
        # Convert the 'Hotel Price Range' column to string type
        df['Hotel Price Range'] = df['Hotel Price Range'].astype(str)

        # Extract numeric values from the 'Hotel Price Range' column using regular expression
        price_range_regex = r'\$(\d+)\s*-\s*\$(\d+)'
        # Extracts numeric value from the 'Hotel Price Range' column using the str.extract() method and converts it to an integer using the astype() method
        df[['Price Range Low', 'Price Range High']] = df['Hotel Price Range'].str.extract(price_range_regex).astype(float)
        df.loc[df['Hotel Price Range'].str.lower().str.contains('none'), ['Price Range Low', 'Price Range High']] = np.nan
        df['Price Range Average'] = df[['Price Range Low', 'Price Range High']].mean(axis=1, skipna=True)
        
        # Create a new column that categorizes the hotel prices into the price ranges using 'pd.cut' function
        df['Price Range Category'] = pd.cut(df['Price Range Average'], 
                                            bins=[0, 100, 200, 300, 400, np.inf], 
                                            labels=['Below $100', '$100-$200', '$201-$300', '$301-$400', 'Above $400'])
        
       
        # Group the DataFrame by the price range categories and calculate the average sustainability for each category
        price_range_avg = df.groupby('Price Range Category')['Hotel Sustainabilty Average'].mean()
        
        # Get the top 5 hotels by sustainability in each price range category
        top_5_by_price_range = pd.DataFrame(columns=['Price Range Category', 'Hotel Name', 'Hotel Sustainabilty Average','Total Number of Reviews', 'Hotel About Section Keywords'])
        for category in price_range_avg.index:
            
            # Get the top 5 hotels by sustainability for the current category
            top_5 = df[df['Price Range Category'] == category].sort_values('Hotel Sustainabilty Average', ascending=False).head(5)
            # Add the category column to the top 5 DataFrame
            top_5['Price Range Category'] = category
            # Append the top 5 DataFrame to the final DataFrame
            top_5_by_price_range = top_5_by_price_range.append(top_5[['Price Range Category', 'Hotel Name', 'Hotel Sustainabilty Average','Total Number of Reviews', 'Hotel About Section Keywords']])
            
        return top_5_by_price_range
    
    def get_top_5_from_all_categories(results):
        all_keywords = []
        for category_keywords in results.values():
            if category_keywords:
                for keyword in category_keywords:
                    all_keywords.append(keyword)
        word_counts = Counter(all_keywords)
        top_5_keywords = sorted(word_counts.items(), key=lambda x: x[1], reverse=True)[:5]
        return top_5_keywords


    def get_top_5_keywords_from_hotel(hotelname, hotels_df):
        categories = ['Enviornmental Keywords', 'Social Keywords', 'Cultural Keywords', 'Economic Keywords', 'Policy Keywords']
        results = {}
        all_keywords = []
        
        for category in categories:
            keywords = hotels_df[hotels_df['Hotel Name'] == hotelname][category].tolist()
            keywords = [word for word in keywords if word != "none"]
            all_keywords.extend(keywords)
            # word_counts = Counter(keywords)
            # results[category] = [word for word, count in word_counts.most_common(5)]
            # top_keywords = word_counts.most_common(5)
            # results[category] = top_keywords[:5]  # Limit the output to top 5 keywords
        # Get top 5 keywords from all categories 
        word_counts = Counter(all_keywords)
        top_keywords = sorted(word_counts.items(), key=itemgetter(1), reverse=True)[:5]
        return top_keywords
    
    # Get the total amount of hotels found in each rating category
    def get_total_hotel_in_each_rating(df):
        # Filters the DataFrame by each rating using boolean indexing and then counts the number of unique hotel names in each group
        hotels_5_star = df[df['Hotel Rating'] == '5']['Hotel Name'].count()
        hotels_4_5_star = df[df['Hotel Rating'] == '4.5']['Hotel Name'].count()
        hotels_4_star = df[df['Hotel Rating'] == '4']['Hotel Name'].count()
        hotels_3_5_star = df[df['Hotel Rating'] == '3.5']['Hotel Name'].count()
        hotels_3_star = df[df['Hotel Rating'] == '3']['Hotel Name'].count()
        hotels_2_5_star = df[df['Hotel Rating'] == '2.5']['Hotel Name'].count()
        hotels_2_star = df[df['Hotel Rating'] == '2']['Hotel Name'].count()
        hotels_1_5_star = df[df['Hotel Rating'] == '1.5']['Hotel Name'].count()
        hotels_1_star = df[df['Hotel Rating'] == '1']['Hotel Name'].count()

        # Returned as a dictionary with the rating as the key and the count as the value.
        return {'5 stars': hotels_5_star, '4.5 stars': hotels_4_5_star, '4 stars': hotels_4_star,
                '3.5 stars': hotels_3_5_star, '3 stars': hotels_3_star, '2.5 stars': hotels_2_5_star,
                '2 stars': hotels_2_star, '1.5 stars': hotels_1_5_star, '1 star': hotels_1_star}


    def update_excel(data_dict, worksheet, skip_columns_after=None):
        # Clear the existing data in the worksheet
        worksheet.delete_rows(1, worksheet.max_row)

        # Write the headers to the first row
        headers = list(data_dict.keys())
        col_index = 1
        for header in headers:
            if header in skip_columns_after:
                worksheet.cell(row=1, column=col_index).value = header
                worksheet.cell(row=1, column=col_index + 1).value = ''
                worksheet.cell(row=1, column=col_index + 2).value = ''
                col_index += 3
            else:
                worksheet.cell(row=1, column=col_index).value = header
                col_index += 1

        # Pad the shorter lists with empty values
        max_length = max([len(data_dict[key]) for key in headers])
        for key, value in data_dict.items():
            if len(value) < max_length:
                data_dict[key] += [''] * (max_length - len(value))

        # Write the data to the remaining rows
        for i in range(max_length):
            col_index = 1
            for header in headers:
                if header in skip_columns_after:
                    worksheet.cell(row=i+2, column=col_index).value = data_dict[header][i]
                    worksheet.cell(row=i+2, column=col_index + 1).value = ''
                    worksheet.cell(row=i+2, column=col_index + 2).value = ''
                    col_index += 3
                else:
                    worksheet.cell(row=i+2, column=col_index).value = data_dict[header][i]
                    col_index += 1

        # Check if there are any invalid XML elements in the workbook
        with open("data_analysis2.xlsx", "rb") as f:
            try:
                etree.parse(f)
            except etree.XMLSyntaxError as e:
                print(f"\nInvalid XML element found in workbook: {e.args[0]}")

        # Save the workbook
        workbook.save("data_analysis2.xlsx")


    
if __name__ == "__main__":
    # Create a dictionary with initially with no data
    research_ques = {'Top 10 Sustainable Hotels':[], 'Hotel About Section Keywords':[] ,'Sustainability Score':[], 'Total Number of Reviews':[],'All Environmental Keywords Found':[], 'All Social Keywords Found':[],
                        'All Environmental Keywords[Environment]':[], 'All Environmental Keywords[Certificate]':[],
                        'All Environmental Keywords[Green Practices]':[], 'All Environmental Keywords[Sustainable Transportation]':[],
                        'All Cultural Keywords Found':[], 'All Economic Keywords Found':[],
                        'All Policy Keywords Found':[], 'Top 10 Enviornmental Keywords':[], 'Top 10 Environmental Keywords[Environment]':[], 'Top 10 Environmental Keywords[Certificate]':[],
                        'Top 10 Environmental Keywords[Green Practices]':[], 'Top 10 Environmental Keywords[Sustainable Transportation]':[],
                        'Top 10 Social Keywords':[], 'Top 10 Cultural Keywords':[], 'Top 10 Economic Keywords' :[], 'Top 10 Policy Keywords':[], 
                        'Top 5 5-star':[],'Top 5 5-star Sustainability Score':[], 'Top 5 5-star Total Number of Reviews':[], 'Top 5 Keywords for 5-star':[], 'Top 5 About Section Keywords for 5-star':[],'Top 5 4.5-star':[], 'Top 5 4.5-star Sustainability Score':[], 
                        'Top 5 4.5-star Total Number of Reviews':[],'Top 5 Keywords for 4.5-star':[],'Top 5 About Section Keywords for 4.5-star':[], 'Top 5 4-star':[],'Top 5 4-star Sustainability Score':[], 'Top 5 4-star Total Number of Reviews':[],'Top 5 Keywords for 4-star':[], 
                        'Top 5 About Section Keywords for 4-star':[], 'Top 5 3.5-star':[], 'Top 5 3.5-star Sustainability Score':[], 'Top 5 3.5-star Total Number of Reviews':[], 'Top 5 Keywords for 3.5-star':[], 'Top 5 About Section Keywords for 3.5-star':[],'Top 5 3-star':[], 
                        'Top 5 3-star Sustainability Score':[], 'Top 5 3-star Total Number of Reviews':[],'Top 5 Keywords for 3-star':[], 'Top 5 About Section Keywords for 3-star':[], 'Top 5 2.5-star':[], 'Top 5 2.5-star Sustainability Score':[], 'Top 5 2.5-star Total Number of Reviews':[], 
                        'Top 5 Keywords for 2.5 star':[],  'Top 5 About Section Keywords for 2.5 star':[], 'Top 5 2-star':[], 'Top 5 2-star Sustainability Score':[],'Top 5 2-star Total Number of Reviews':[], 'Top 5 Keywords for 2-star':[], 'Top 5 About Section Keywords for 2-star':[],
                        'Top 5 1.5-star':[],'Top 5 1.5-star Sustainability Score':[], 'Top 5 1.5-star Total Number of Reviews':[], 'Top 5 Keywords for 1.5-star':[], 'Top 5 About Section Keywords for 1.5-star':[], 'Top 5 1-star':[], 'Top 5 1-star Sustainability Score':[], 'Top 5 1-star Total Number of Reviews':[],'Top 5 Keywords for 1-star':[],
                        'Top 5 About Section Keywords for 1-star':[],


                        'Top 5 Below $100': [], 'Top 5 Below $100 Sustainability Score':[], 'Top 5 Below $100 Total Number of Reviews':[],'Top 5 Keywords for Below $100':[], 'Top 5 About Section Keywords for Below $100':[],
                        'Top 5 $100-$200': [], 'Top 5 $100-$200 Sustainability Score': [], 'Top 5 $100-$200 Total Number of Reviews':[], 'Top 5 Keywords for $100-$200':[], 'Top 5 About Section Keywords for $100-$200':[], 'Top 5 $201-$300':[],'Top 5 $201-$300 Sustainability Score':[], 'Top 5 $201-$200 Total Number of Reviews':[], 
                        'Top 5 Keywords for $201-$300':[], 'Top 5 About Section Keywords for $201-$300':[], 'Top 5 $301-$400':[],'Top 5 $301-$400 Sustainability Score':[], 'Top 5 $301-$400 Total Number of Reviews':[], 'Top 5 Keywords for $301-$400':[], 'Top 5 About Section Keywords for $301-$400':[], 'Top 5 Above $400':[], 
                        'Top 5 Above $400 Sustainability Score':[], 'Top 5 Above $400 Total Number of Reviews':[],'Top 5 Keywords for Above $400' :[], 'Top 5 About Section Keywords for Above $400' :[], 'Total Number of Hotels 5-star Rating':[], 
                        'Total Number of Hotels 4.5-star Rating' : [], 'Total Number of Hotels 4-star Rating' :[], 'Total Number of Hotels 3.5-star Rating':[],
                        'Total Number of Hotels 3-star Rating':[], 'Total Number of Hotels 2.5-star Rating':[], 'Total Number of Hotels 2-star Rating':[],
                        'Total Number of Hotels 1.5-star Rating':[], 'Total Number of Hotels 1-star Rating':[], 
                        }

    # Get top 10 hotels with highest sustainabilty score 
    df = pd.read_excel('data_analysis.xlsx', usecols=['Hotel Name', 'Hotel About Section Keywords','Hotel Sustainabilty Average', 'Total Number of Reviews', 'Enviornmental Keywords', 
                                                      'Environmental Keywords[Environment]', 'Environmental Keywords[Certificate]', 'Environmental Keywords[Green Practices]',
                                                      'Environmental Keywords[Sustainable Transportation]',
                                                      'Social Keywords', 'Cultural Keywords','Economic Keywords', 'Policy Keywords'])
    # Filter the data frame based on the condition
    df = df[df['Total Number of Reviews'] > 100]

    print("Processing...top 10")
    top_10 = Analysis.get_top_10(df, 'Hotel Sustainabilty Average')
    top_10_hotel_names = top_10['Hotel Name'].tolist()
    top_10_hotel_about_section_keywords = top_10['Hotel About Section Keywords'].tolist()
    top_10_hotel_scores = top_10['Hotel Sustainabilty Average'].tolist()
    top_10_hotel_total_num_reviews = top_10['Total Number of Reviews'].tolist()

    # Append top 10 hotels with highest sustainabilty score to dictionary 
    for hotel_name, hotel_about_section_keywords, hotel_score, hotel_total_review in zip(top_10_hotel_names, top_10_hotel_about_section_keywords,top_10_hotel_scores, top_10_hotel_total_num_reviews):
        research_ques['Top 10 Sustainable Hotels'].append(hotel_name)
        research_ques['Hotel About Section Keywords'].append(hotel_about_section_keywords)
        research_ques['Sustainability Score'].append(hotel_score)
        research_ques['Total Number of Reviews'].append(hotel_total_review)

    # Get all review keywords from top 10 hotels
    top_10_with_keywords = Analysis.get_hotel_keywords(df,top_10)

    # For each sustainability category append the keyword found for each hotel    
    for hotel, keywords in top_10_with_keywords:
        # print(hotel)
        # Set so results do not contain duplicates before appending to 'research_ques' dict 
        enviornmental = []
        enviornment = []
        certificate = []
        green_practices = []
        sustainable_transportation = []
        social = []
        cultural = []
        economic = []
        policy = []
        # For each hotel find the keywords and append them to specific list depending on the category
        for keyword, category in keywords:
            # Normalize keyword by converting to lowercase, removing punctuation, and splitting on commas
            keyword = keyword.lower().strip('.,;:')
            keywords_list = [k.strip() for k in keyword.split(',')]
            for k in keywords_list:
                if category == 'Environmental':
                    enviornmental.append(k)
                if category == '[Environment]':
                    enviornment.append(k)
                if category == '[Certificate]':
                    certificate.append(k)
                if category == '[Green Practices]':
                    green_practices.append(k)
                if category == '[Sustainable Transportation]':
                    sustainable_transportation.append(k)
                if category == 'Social':
                    social.append(k)
                if category == 'Cultural':
                    cultural.append(k)
                if category == 'Economic':
                    economic.append(k)
                if category == 'Policy':
                    policy.append(k)
            # print(f"\t{category}: {k}")

        # Count the duplicates in each sustainability category list
        environmental_counts = Counter(enviornmental)
        social_counts = Counter(social)
        cultural_counts = Counter(cultural)
        economic_counts = Counter(economic)
        policy_counts = Counter(policy)
        enviornment_counts = Counter(enviornment)
        certificate_counts = Counter(certificate)
        green_practices_counts = Counter(green_practices)
        sustainable_transportation_counts = Counter(sustainable_transportation)
        
        # Append count data for each category to corresponding key in dictionary
        research_ques['All Environmental Keywords Found'].append(", ".join([f"{k}: {v}" for k, v in environmental_counts.items()]))
        research_ques['All Environmental Keywords[Environment]'].append(", ".join([f"{k}: {v}" for k, v in enviornment_counts.items()]))
        research_ques['All Environmental Keywords[Certificate]'].append(", ".join([f"{k}: {v}" for k, v in certificate_counts.items()]))
        research_ques['All Environmental Keywords[Green Practices]'].append(", ".join([f"{k}: {v}" for k, v in green_practices_counts.items()]))
        research_ques['All Environmental Keywords[Sustainable Transportation]'].append(", ".join([f"{k}: {v}" for k, v in sustainable_transportation_counts.items()]))
        research_ques['All Social Keywords Found'].append(", ".join([f"{k}: {v}" for k, v in social_counts.items()]))
        research_ques['All Cultural Keywords Found'].append(", ".join([f"{k}: {v}" for k, v in cultural_counts.items()]))
        research_ques['All Economic Keywords Found'].append(", ".join([f"{k}: {v}" for k, v in economic_counts.items()]))
        research_ques['All Policy Keywords Found'].append(", ".join([f"{k}: {v}" for k, v in policy_counts.items()]))
        enviornmental.clear()
        enviornment.clear()
        certificate.clear()
        green_practices.clear()
        sustainable_transportation.clear()
        social.clear()
        cultural.clear()
        economic.clear()
        policy.clear()


    # Get the top 10 keywords from each category 
    df2 = pd.read_excel('data_analysis.xlsx', usecols=['Enviornmental Keywords', 'Social Keywords', 'Cultural Keywords',
                                                      'Economic Keywords','Environmental Keywords[Environment]', 'Environmental Keywords[Certificate]', 'Environmental Keywords[Green Practices]',
                                                      'Environmental Keywords[Sustainable Transportation]', 'Policy Keywords'])
    print("Processing...top enviornmental keywords")
    top_10_enviornmental = Analysis.get_top_10_keywords(df2,'Enviornmental Keywords')
    top_10_enviornmental_list = [str(item) for item in top_10_enviornmental]  # convert tuple to list of strings

    print("Processing...top enviornmental keywords subcatergories")
    top_10_enviornment = Analysis.get_top_10_keywords(df2,'Environmental Keywords[Environment]')
    top_10_enviornment_list = [str(item) for item in  top_10_enviornment]  # convert tuple to list of strings
    top_10_certificate = Analysis.get_top_10_keywords(df2,'Environmental Keywords[Certificate]')
    top_10_certficate_list = [str(item) for item in  top_10_certificate]  # convert tuple to list of strings
    top_10_green_practices = Analysis.get_top_10_keywords(df2,'Environmental Keywords[Green Practices]')
    top_10_green_practices_list = [str(item) for item in top_10_green_practices]  # convert tuple to list of strings
    top_10_sustainable_transportation = Analysis.get_top_10_keywords(df2,'Environmental Keywords[Sustainable Transportation]')
    top_10_sustainable_transportation_list  = [str(item) for item in top_10_sustainable_transportation]  # convert tuple to list of strings

    print("Processing...top social keywords")
    top_10_social = Analysis.get_top_10_keywords(df2,'Social Keywords')
    top_10_social_list = [str(item) for item in top_10_social]  # convert tuple to list of strings
    print("Processing...top cultural keywords")
    top_10_cultural = Analysis.get_top_10_keywords(df2,'Cultural Keywords')
    top_10_cultural_list = [str(item) for item in top_10_cultural]  # convert tuple to list of strings
    print("Processing...top economic keywords") 
    top_10_economic = Analysis.get_top_10_keywords(df2,'Economic Keywords')
    top_10_economic_list = [str(item) for item in top_10_economic]  # convert tuple to list of strings
    print("Processing...top policy keywords") 
    top_10_policy = Analysis.get_top_10_keywords(df2,'Policy Keywords')
    top_10_policy_list = [str(item) for item in top_10_policy]  # convert tuple to list of strings

    print('Writing to dic')
    # Append top 10 keywords from each category to dictionary
    # itertools.zip_longest() function to iterate over all lists, even if they have different lengths, value of "" to fill any missing values.
    for enviornmental_keywords, enviornment_keywords, certificate_keywords, green_practices_keywords, sustainable_transportation_keywords, social_keywords, cultural_keywords, economic_keywords, policy_keywords in itertools.zip_longest(
            top_10_enviornmental_list, top_10_enviornment_list, top_10_certficate_list,  top_10_green_practices_list, top_10_sustainable_transportation_list, top_10_social_list, top_10_cultural_list, top_10_economic_list, top_10_policy_list,
            fillvalue=""):
        research_ques['Top 10 Enviornmental Keywords'].append(enviornmental_keywords)
        research_ques['Top 10 Environmental Keywords[Environment]'].append(enviornment_keywords)
        research_ques['Top 10 Environmental Keywords[Certificate]'].append(certificate_keywords)
        research_ques['Top 10 Environmental Keywords[Green Practices]'].append(green_practices_keywords)
        research_ques['Top 10 Environmental Keywords[Sustainable Transportation]'].append(sustainable_transportation_keywords)        
        research_ques['Top 10 Social Keywords'].append(social_keywords)
        research_ques['Top 10 Cultural Keywords'].append(cultural_keywords)
        research_ques['Top 10 Economic Keywords'].append(economic_keywords)
        research_ques['Top 10 Policy Keywords'].append(policy_keywords)
    
    hotels_df = pd.read_excel('data_analysis.xlsx', usecols=['Hotel Name', 'Enviornmental Keywords','Social Keywords', 'Cultural Keywords','Economic Keywords', 'Policy Keywords'])

    df3 = pd.read_excel('data_analysis.xlsx', usecols=['Hotel Name','Hotel Rating', 'Hotel Sustainabilty Average', 'Total Number of Reviews', 'Hotel About Section Keywords'])
    # Create a data frame without duplicates of hotel name
    df3 = df3.drop_duplicates(subset=['Hotel Name'])
    # Filter the data frame based on the condition
    df3 = df3[df3['Total Number of Reviews'] > 100]

    print("Finding top 5 hotels in each rating...")
    # Results of the top 5 hotels in each rating based on 'Hotel Sustainabilty Average'
    results_dict = Analysis.get_top_5_stars(df3)
    print(results_dict)

    # Append top 5 hotels from each rating category to dictionary
    for rating, hotels in results_dict.items():
        print(f'Top 5 most sustainable hotels with {rating} star rating:')
    
        for hotel in hotels:
            if rating == '5':
                research_ques['Top 5 5-star'].append(hotel[0])
                research_ques['Top 5 5-star Sustainability Score'].append(hotel[1])
                research_ques['Top 5 5-star Total Number of Reviews'].append(hotel[2])
                top_5_keywords_from_hotel = Analysis.get_top_5_keywords_from_hotel(hotel[0], hotels_df)
                top_5_keywords_str = []
                for keyword, frequency in top_5_keywords_from_hotel:
                    print(f"Keyword: {keyword}, Frequency: {frequency}")
                    top_5_keywords_str.append(keyword + " (" + str(frequency) + ")")
                
                research_ques['Top 5 Keywords for 5-star'].append(", ".join(top_5_keywords_str))
                research_ques['Top 5 About Section Keywords for 5-star'].append(hotel[3])

            elif rating == '4.5':
                research_ques['Top 5 4.5-star'].append(hotel[0])
                research_ques['Top 5 4.5-star Sustainability Score'].append(hotel[1])
                research_ques['Top 5 4.5-star Total Number of Reviews'].append(hotel[2])

                top_5_keywords_from_hotel = Analysis.get_top_5_keywords_from_hotel(hotel[0], hotels_df)
                top_5_keywords_str = []
                for keyword, frequency in top_5_keywords_from_hotel:
                    print(f"Keyword: {keyword}, Frequency: {frequency}")
                    top_5_keywords_str.append(keyword + " (" + str(frequency) + ")")
                        
                research_ques['Top 5 Keywords for 4.5-star'].append(", ".join(top_5_keywords_str))
                research_ques['Top 5 About Section Keywords for 4.5-star'].append(hotel[3])

            elif rating == '4':
                research_ques['Top 5 4-star'].append(hotel[0])
                research_ques['Top 5 4-star Sustainability Score'].append(hotel[1])
                research_ques['Top 5 4-star Total Number of Reviews'].append(hotel[2])

                top_5_keywords_from_hotel = Analysis.get_top_5_keywords_from_hotel(hotel[0], hotels_df)
                top_5_keywords_str = []
                for keyword, frequency in top_5_keywords_from_hotel:
                    print(f"Keyword: {keyword}, Frequency: {frequency}")
                    top_5_keywords_str.append(keyword + " (" + str(frequency) + ")")
                research_ques['Top 5 Keywords for 4-star'].append(", ".join(top_5_keywords_str))
                research_ques['Top 5 About Section Keywords for 4-star'].append(hotel[3])

            elif rating == '3.5':
                research_ques['Top 5 3.5-star'].append(hotel[0])
                research_ques['Top 5 3.5-star Sustainability Score'].append(hotel[1])
                research_ques['Top 5 3.5-star Total Number of Reviews'].append(hotel[2])

                top_5_keywords_from_hotel = Analysis.get_top_5_keywords_from_hotel(hotel[0], hotels_df)
                top_5_keywords_str = []
                for keyword, frequency in top_5_keywords_from_hotel:
                    print(f"Keyword: {keyword}, Frequency: {frequency}")
                    top_5_keywords_str.append(keyword + " (" + str(frequency) + ")")

                research_ques['Top 5 Keywords for 3.5-star'].append(", ".join(top_5_keywords_str))
                research_ques['Top 5 About Section Keywords for 3.5-star'].append(hotel[3])
   
            elif rating == '3':
                research_ques['Top 5 3-star'].append(hotel[0])
                research_ques['Top 5 3-star Sustainability Score'].append(hotel[1])
                research_ques['Top 5 3-star Total Number of Reviews'].append(hotel[2])

                top_5_keywords_from_hotel = Analysis.get_top_5_keywords_from_hotel(hotel[0], hotels_df)
                top_5_keywords_str = []
                for keyword, frequency in top_5_keywords_from_hotel:
                    print(f"Keyword: {keyword}, Frequency: {frequency}")
                    top_5_keywords_str.append(keyword + " (" + str(frequency) + ")")
                research_ques['Top 5 Keywords for 3-star'].append(", ".join(top_5_keywords_str))
                research_ques['Top 5 About Section Keywords for 3-star'].append(hotel[3])
   
            elif rating == '2.5':
                research_ques['Top 5 2.5-star'].append(hotel[0])
                research_ques['Top 5 2.5-star Sustainability Score'].append(hotel[1])
                research_ques['Top 5 2.5-star Total Number of Reviews'].append(hotel[2])

                top_5_keywords_from_hotel = Analysis.get_top_5_keywords_from_hotel(hotel[0], hotels_df)
                top_5_keywords_str = []
                for keyword, frequency in top_5_keywords_from_hotel:
                    print(f"Keyword: {keyword}, Frequency: {frequency}")
                    top_5_keywords_str.append(keyword + " (" + str(frequency) + ")")
                research_ques['Top 5 Keywords for 2.5 star'].append(", ".join(top_5_keywords_str))
                research_ques['Top 5 About Section Keywords for 2.5 star'].append(hotel[3])

            elif rating == '2':
                research_ques['Top 5 2-star'].append(hotel[0])
                research_ques['Top 5 2-star Sustainability Score'].append(hotel[1])
                research_ques['Top 5 2-star Total Number of Reviews'].append(hotel[2])

                top_5_keywords_from_hotel = Analysis.get_top_5_keywords_from_hotel(hotel[0], hotels_df)
                top_5_keywords_str = []
                for keyword, frequency in top_5_keywords_from_hotel:
                    print(f"Keyword: {keyword}, Frequency: {frequency}")
                    top_5_keywords_str.append(keyword + " (" + str(frequency) + ")")
                research_ques['Top 5 Keywords for 2-star'].append(", ".join(top_5_keywords_str))
                research_ques['Top 5 About Section Keywords for 2-star'].append(hotel[3])
  
            elif rating == '1.5':
                research_ques['Top 5 1.5-star'].append(hotel[0])
                research_ques['Top 5 1.5-star Sustainability Score'].append(hotel[1])
                research_ques['Top 5 1.5-star Total Number of Reviews'].append(hotel[2])

                top_5_keywords_from_hotel = Analysis.get_top_5_keywords_from_hotel(hotel[0], hotels_df)
                top_5_keywords_str = []
                for keyword, frequency in top_5_keywords_from_hotel:
                    print(f"Keyword: {keyword}, Frequency: {frequency}")
                    top_5_keywords_str.append(keyword + " (" + str(frequency) + ")")
                research_ques['Top 5 Keywords for 1.5-star'].append(", ".join(top_5_keywords_str))
                research_ques['Top 5 About Section Keywords for 1.5-star'].append(hotel[3])

            elif rating == '1':
                research_ques['Top 5 1-star'].append(hotel[0])
                research_ques['Top 5 1-star Sustainability Score'].append(hotel[1])
                research_ques['Top 5 1-star Total Number of Reviews'].append(hotel[2])

                top_5_keywords_from_hotel = Analysis.get_top_5_keywords_from_hotel(hotel[0], hotels_df)
                top_5_keywords_str = []
                for keyword, frequency in top_5_keywords_from_hotel:
                    print(f"Keyword: {keyword}, Frequency: {frequency}")
                    top_5_keywords_str.append(keyword + " (" + str(frequency) + ")")
                research_ques['Top 5 Keywords for 1-star'].append(", ".join(top_5_keywords_str))
                research_ques['Top 5 About Section Keywords for 1-star'].append(hotel[3])
            print(f'{hotel[0]}: {hotel[1]}')

    df4 = pd.read_excel('data_analysis.xlsx', usecols=['Hotel Name','Hotel Price Range', 'Hotel Sustainabilty Average', 'Total Number of Reviews', 'Hotel About Section Keywords'])
    # Create a data frame without duplicates of hotel name
    df4= df4.drop_duplicates(subset=['Hotel Name'])
    # Filter the data frame based on the condition
    df4 = df4[df4['Total Number of Reviews'] > 100]

    print("Finding top 5 hotels in each price range...")
    # Results of the top 5 hotels in each price range based on 'Hotel Sustainabilty Average'
    top_5_hotels_by_price_range = Analysis.get_top_5_by_price_range(df4)
    print(top_5_hotels_by_price_range)

    total_price_range = Analysis.get_total_price_range(df4)
    print("TOTAL PRICE RANGE:" ,total_price_range)

    hotels_df = pd.read_excel('data_analysis.xlsx', usecols=['Hotel Name', 'Enviornmental Keywords','Social Keywords', 'Cultural Keywords','Economic Keywords', 'Policy Keywords','Total Number of Reviews', 'Hotel About Section Keywords'])
    print(hotels_df.columns)
    # loop through each price range category and add the top 5 hotels to the dictionary
    for category in top_5_hotels_by_price_range['Price Range Category'].unique():
        hotels = top_5_hotels_by_price_range[top_5_hotels_by_price_range['Price Range Category'] == category][['Hotel Name', 'Hotel Sustainabilty Average','Total Number of Reviews', 'Hotel About Section Keywords']]
        for index, row in hotels.iterrows():
            if category == 'Below $100':
                research_ques['Top 5 Below $100'].append(row['Hotel Name'])
                research_ques['Top 5 Below $100 Sustainability Score'].append(row['Hotel Sustainabilty Average'])
                research_ques['Top 5 Below $100 Total Number of Reviews'].append(row['Total Number of Reviews'])

                top_5_keywords_from_hotel = Analysis.get_top_5_keywords_from_hotel(row['Hotel Name'], hotels_df)
                top_5_keywords_str = []
                for keyword, frequency in top_5_keywords_from_hotel:
                    print(f"Keyword: {keyword}, Frequency: {frequency}")
                    top_5_keywords_str.append(keyword + " (" + str(frequency) + ")")
                research_ques['Top 5 Keywords for Below $100'].append(", ".join(top_5_keywords_str))
                research_ques['Top 5 About Section Keywords for Below $100'].append(row['Hotel About Section Keywords'])

            if category == '$100-$200':
                research_ques['Top 5 $100-$200'].append(row['Hotel Name'])
                research_ques['Top 5 $100-$200 Sustainability Score'].append(row['Hotel Sustainabilty Average'])
                research_ques['Top 5 $100-$200 Total Number of Reviews'].append(row['Total Number of Reviews'])

                top_5_keywords_from_hotel = Analysis.get_top_5_keywords_from_hotel(row['Hotel Name'], hotels_df)
                top_5_keywords_str = []
                for keyword, frequency in top_5_keywords_from_hotel:
                    print(f"Keyword: {keyword}, Frequency: {frequency}")
                    top_5_keywords_str.append(keyword + " (" + str(frequency) + ")")
                research_ques['Top 5 Keywords for $100-$200'].append(", ".join(top_5_keywords_str))
                research_ques['Top 5 About Section Keywords for $100-$200'].append(row['Hotel About Section Keywords'])

            if category == '$201-$300':
                research_ques['Top 5 $201-$300'].append(row['Hotel Name'])
                research_ques['Top 5 $201-$300 Sustainability Score'].append(row['Hotel Sustainabilty Average'])
                research_ques['Top 5 $201-$200 Total Number of Reviews'].append(row['Total Number of Reviews'])

                top_5_keywords_from_hotel = Analysis.get_top_5_keywords_from_hotel(row['Hotel Name'], hotels_df)
                top_5_keywords_str = []
                for keyword, frequency in top_5_keywords_from_hotel:
                    print(f"Keyword: {keyword}, Frequency: {frequency}")
                    top_5_keywords_str.append(keyword + " (" + str(frequency) + ")")
                research_ques['Top 5 Keywords for $201-$300'].append(", ".join(top_5_keywords_str))
                research_ques['Top 5 About Section Keywords for $201-$300'].append(row['Hotel About Section Keywords'])

            if category == '$301-$400':
                research_ques['Top 5 $301-$400'].append(row['Hotel Name'])
                research_ques['Top 5 $301-$400 Sustainability Score'].append(row['Hotel Sustainabilty Average'])
                research_ques['Top 5 $301-$400 Total Number of Reviews'].append(row['Total Number of Reviews'])

                top_5_keywords_from_hotel = Analysis.get_top_5_keywords_from_hotel(row['Hotel Name'], hotels_df)
                top_5_keywords_str = []
                for keyword, frequency in top_5_keywords_from_hotel:
                    print(f"Keyword: {keyword}, Frequency: {frequency}")
                    top_5_keywords_str.append(keyword + " (" + str(frequency) + ")")
                research_ques['Top 5 Keywords for $301-$400'].append(", ".join(top_5_keywords_str))
                research_ques['Top 5 About Section Keywords for $301-$400'].append(row['Hotel About Section Keywords'])

            if category == 'Above $400':
                research_ques['Top 5 Above $400'].append(row['Hotel Name'])
                research_ques['Top 5 Above $400 Sustainability Score'].append(row['Hotel Sustainabilty Average'])
                research_ques['Top 5 Above $400 Total Number of Reviews'].append(row['Total Number of Reviews'])

                top_5_keywords_from_hotel = Analysis.get_top_5_keywords_from_hotel(row['Hotel Name'], hotels_df)
                top_5_keywords_str = []
                for keyword, frequency in top_5_keywords_from_hotel:
                    print(f"Keyword: {keyword}, Frequency: {frequency}")
                    top_5_keywords_str.append(keyword + " (" + str(frequency) + ")")
                research_ques['Top 5 Keywords for Above $400'].append(", ".join(top_5_keywords_str))
                research_ques['Top 5 About Section Keywords for Above $400'].append(row['Hotel About Section Keywords'])


    df5 = pd.read_excel('data_analysis.xlsx', usecols=['Hotel Name','Hotel Rating', 'Total Number of Reviews'])
    # Create a data frame without duplicates of hotel name
    df5 = df5.drop_duplicates(subset=['Hotel Name'])
    # Get the total number of hotels found in each rating 
    total_hotel_in_each_rating = Analysis.get_total_hotel_in_each_rating(df5)
    print(total_hotel_in_each_rating)
    for rating, total_num_of_hotel in total_hotel_in_each_rating.items():
        if rating == '5 stars':
            research_ques['Total Number of Hotels 5-star Rating'].append(total_num_of_hotel)
        if rating == '4.5 stars':
            research_ques['Total Number of Hotels 4.5-star Rating'].append(total_num_of_hotel)
        if rating == '4 stars':
            research_ques['Total Number of Hotels 4-star Rating'].append(total_num_of_hotel)
        if rating == '3.5 stars':
            research_ques['Total Number of Hotels 3.5-star Rating'].append(total_num_of_hotel)
        if rating == '3 stars':
            research_ques['Total Number of Hotels 3-star Rating'].append(total_num_of_hotel)
        if rating == '2.5 stars':
            research_ques['Total Number of Hotels 2.5-star Rating'].append(total_num_of_hotel)
        if rating == '2 stars':
            research_ques['Total Number of Hotels 2-star Rating'].append(total_num_of_hotel)
        if rating == '1.5 stars':
            research_ques['Total Number of Hotels 1.5-star Rating'].append(total_num_of_hotel)
        if rating == '1 star':
            research_ques['Total Number of Hotels 1-star Rating'].append(total_num_of_hotel)


    
    # # Update results to excel file
    Analysis.update_excel(research_ques, worksheet, skip_columns_after=['All Policy Keywords Found', 'Top 10 Policy Keywords',
    'Top 5 About Section Keywords for 5-star', 'Top 5 About Section Keywords for 4.5-star', 'Top 5 About Section Keywords for 4-star', 'Top 5 About Section Keywords for 3.5-star',
    'Top 5 About Section Keywords for 3-star', 'Top 5 About Section Keywords for 2.5 star', 'Top 5 About Section Keywords for 2-star', 'Top 5 About Section Keywords for 1.5-star', 'Top 5 About Section Keywords for 1-star',
    'Top 5 About Section Keywords for Below $100', 'Top 5 About Section Keywords for $100-$200', 'Top 5 About Section Keywords for $201-$300', 'Top 5 About Section Keywords for $301-$400', 'Top 5 About Section Keywords for Above $400',
    ])


