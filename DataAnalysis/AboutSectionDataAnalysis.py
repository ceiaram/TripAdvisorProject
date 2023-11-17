# Import the Counter class from the collections module
from collections import Counter
import openpyxl
import pymongo
import dns # required for connecting with SRV
import nltk
import statistics
import itertools 
from lxml import etree
import pandas as pd
import re

# Create a new workbook
workbook = openpyxl.Workbook()
# Select the active worksheet
worksheet = workbook.active

def findKeywords(text, keywords, str_list_name):
       word_variations = []
       #store the two-worded keywords and their respective values
       two_worded_keywords = {}

       # Includes two-worded keywords in the dictionary and adds their value to the score if both words are found in the text
       for keyword in keywords:
        words = keyword.split()
        if len(words) > 1:
            first_word = words[0]
            second_word = words[1]
            word_variations.append(r'\b{}\w*\b.*?\b{}\w*\b|\b{}\w*\b \w*\b'.format(re.escape(keyword), re.escape(second_word), re.escape(first_word)))
            two_worded_keywords[keyword] = keywords[keyword]

       # Pattern matches either a single word from the dictionary keys, or a two-word phrase that has been split into its constituent words and combined with the \w*
       pattern = re.compile(r'\b(?:{})\w*\b|\b(?:{})\b'.format('|'.join(map(re.escape, keywords.keys())), '|'.join(word_variations)), re.IGNORECASE)

       # Scores for each category 
       score = 0

       # search for a keyword in the text from pattern
       match = pattern.findall(text)

       matchesFound = []  
       matchesFound.extend(match)

       # remove duplicates from the list of keyword matches
       matchesFound = list(set(matchesFound))

       printed_matches = set()

       # Loop over all matches found and add correct values based on keywords
       for match in matchesFound:
            # Check if the match is a one-worded keyword and add its value to the score
            for keyword, value in keywords.items():
                if keyword in match.lower() and match not in printed_matches:
                    score += value
                    print(f"Matched keyword '{keyword}' with value {value}: '{match}'")
                    printed_matches.add(match)
            # Check if the match is a two-worded keyword and add its value to the score 
            for keyword, value in two_worded_keywords.items():
                if keyword in match.lower() and match not in printed_matches:
                    score += value
                    print(f"Matched keyword '{keyword}' with value {value}: '{match}'")
                    printed_matches.add(match)

       if printed_matches:
        print("Keywords found: ", printed_matches)
        print("Score: ", score)
       else:
        print('No keywords found.')
        printed_matches.add('none')

       return printed_matches

def update_excel(data_dict, worksheet, skip_columns_after=None):
        # Clear the existing data in the worksheet
        worksheet.delete_rows(1, worksheet.max_row)

        # Write the headers to the first row
        headers = list(data_dict.keys())
        col_index = 1
        for header in headers:
            if header in skip_columns_after:
                worksheet.cell(row=1, column=col_index).value = header
                worksheet.cell(row=1, column=col_index + 1).value = ''
                worksheet.cell(row=1, column=col_index + 2).value = ''
                col_index += 3
            else:
                worksheet.cell(row=1, column=col_index).value = header
                col_index += 1

        # Pad the shorter lists with empty values
        max_length = max([len(data_dict[key]) for key in headers])
        for key, value in data_dict.items():
            if len(value) < max_length:
                data_dict[key] += [''] * (max_length - len(value))

        # Write the data to the remaining rows
        for i in range(max_length):
            col_index = 1
            for header in headers:
                if header in skip_columns_after:
                    worksheet.cell(row=i+2, column=col_index).value = data_dict[header][i]
                    worksheet.cell(row=i+2, column=col_index + 1).value = ''
                    worksheet.cell(row=i+2, column=col_index + 2).value = ''
                    col_index += 3
                else:
                    worksheet.cell(row=i+2, column=col_index).value = data_dict[header][i]
                    col_index += 1

        # Check if there are any invalid XML elements in the workbook
        with open("about_section_keywords_analysis.xlsx", "rb") as f:
            try:
                etree.parse(f)
            except etree.XMLSyntaxError as e:
                print(f"\nInvalid XML element found in workbook: {e.args[0]}")

        # Save the workbook
        workbook.save("about_section_keywords_analysis.xlsx")


#  HAVE ALL KEYWORDS LOWERCASE TO FIND SCORES
environmental_keywords = {
    'automatic light': 1,   # TWO-WORDED KEYWORDS MUST COME BEFORE ONE-WORDED KEY WORDS THAT ARE FOUND IN THE TWO-WORDED KEYWORDS OTHERWIS IT WON'T PICK UP THE TWO-WORDED KEYWORDS/CCORRECT SCORES 
    'clean air': 1,  
    'clean environment': 1, 
    'energy sav': 1, ##
    'environmental certificat': 1, 
    'environmental concern': 1, 
    'environmental manage': 1, 
    'environmentally friendly': 1, 
    'ev charg': 1, 
    'ev park': 1, 
    'electronic car': 1, 
    'electric car': 1, 
    'iso 14001': 1, 
    'towel reuse': 1, 
    'linen reuse' : 1, 
    'led light': 1, 
    'leed certificat': 1, 
    'locally grown': 1, 
    'motion sensor': 1, 
    'green space': 1, 
    'green globe': 1, 
    'green hotel': 1, 
    'green issue': 1, 
    'green lodge': 1, 
    'green practice': 1, 
    'green certificat': 1, 
    'recharge station': 1, 
    'sustainable material': 1, 
    'sustainable land': 1, 
    'sustainable transportation': 1, 
    'waste conservat': 1, 
    'water sav': 1, 
    'vehicle charg': 1, 
    # 'green': 1,
    'conservat': 1, 
    'ecological ': 1, 
    'eco-friendly': 1, 
    # 'environment': 1, 
    # 'reuse': 1,
    'recycle': 1, 
    'renewable': 1, 
    'solar': 1, 
    'organic': 1, 
    'biodegrad': 1, 
    'biodiversit': 1, 
    'restorat': 1, 
    'hybrid': 1, 
    'natur': 1, 
    'efficient': 1, 
    'vegetarian': 1, 
    'vegan': 1, 
    'healthy': 1, 
    'sustainability': 1, 
    # 'leed': 1,
    'non-fuel': 1, 
    'non fuel' : 1, 
    'reduce': 1 
    }

environmental_keywords_environment = {
        'clean air': 1,  
        'clean environment': 1, 
        'green space': 1, 
        'green globe': 1, 
        'green hotel': 1, 
        'green issue': 1, 
        'green lodge': 1, 
        'green practice': 1, 
        'sustainable land': 1,
        'ecological ': 1, 
        'sustainability': 1,
    }

environmental_keywords_certificate = {
        'environmental certificat': 1,
        'leed certificat': 1,
        'green certificat': 1,
        'iso 14001': 1,
    }

environmental_keywords_practices = {
        'automatic light': 1,
        'environmental concern': 1, 
        'environmentally friendly': 1,
        'environmental manage': 1,
        'towel reuse': 1, 
        'led light': 1,
        'linen reuse' : 1,
        'locally grown': 1,
        'energy sav': 1,
        'motion sensor': 1,
        'sustainable material': 1,
        'waste conservat': 1,
        'water sav': 1,
        'eco-friendly': 1,
        'recycle': 1,
        'renewable': 1,
        'conservat': 1,
        'solar': 1,
        'organic': 1,
        'biodegrad': 1,
        'biodiversit': 1,
        'restorat': 1,
        'natur': 1,
        'efficient': 1,
        'vegetarian': 1,
        'vegan': 1,
        'healthy': 1,
        'reduce': 1
    }

environmental_keywords_sustainable_transportation = {
        'ev charg': 1,
        'ev park': 1,
        'electronic car': 1,
        'electric car': 1,
        'sustainable transportation': 1,
        'recharge station': 1,
        'non fuel' : 1,
        'vehicle charg': 1,
        'non-fuel': 1,
        'hybrid': 1,
    }


social_keywords = {
    'corporate responsibilit': 1,
    'community support': 1,
    'ethical practice': 1,
    'local employee': 1,
    'local purchase': 1,
    'local communit': 1,
    'responsible consumer behavior': 1,
    'responsible customer behavior': 1,
    'social responsibilit': 1,
    'sustainable communit': 1,
    # 'communit': 1,
    # 'responsibl': 1,
    # 'involvement': 1,
    }

cultural_keywords = {
    'cultural sustainabilit': 1,
    'safe work': 1,
    'work safe' : 1,
    'heritage': 1,
    'inclusi': 1,
    'traditional': 1,
    'diversit': 1,
    'equit': 1,
    # 'fairly': 1,
    'fairness' : 1,
    'dei': 1,
    }

economic_keywords = {
    'economical' : 1,
    'market demand': 1,
    'competitive': 1,
    'advantage': 1,
    'innovat': 1,
    'profit': 1,
    'invest': 1,
    'value': 1,
    # 'renovat': 1
    }

policy_keywords = {
    'corporate social responsibilit': 1,
    'environmental regulation': 1,
    'labor regulat': 1,
    'organizational identit': 1,
    'social responsibilit': 1,
    'social justice': 1,
    'tax polic': 1,
    'zoning polic': 1,
    'stakeholder': 1,
    'regulation': 1,
    'policy': 1, 
    'policies': 1,
    'transparen': 1   
    }


def get_top_10_keywords(df, col):
        top_10_keywords = df[col].tolist()
        # Exclude 'none', so it won't be counted as a common word
        top_10_keywords = [word for word in top_10_keywords if word != "none"]
        word_counts = Counter(top_10_keywords)

        result = []
        for word, count in word_counts.most_common(10):
            result.append((word, count))
        return result

research_ques2 = {'Top 10 Environmental Keywords[Environment]' : [], 'Top 10 Environmental Keywords[Certificate]':[], 'Top 10 Environmental Keywords[Green Practices]':[],
                 'Top 10 Environmental Keywords[Sustainable Transportation]' :[], 'Top 10 Social Keywords':[], 'Top 10 Cultural Keywords':[], 'Top 10 Economic Keywords':[],
                 'Top 10 Policy Keywords':[]}
# Get the top 10 keywords from each category 
df2 = pd.read_excel('about_section_keywords.xlsx', usecols=['Environmental Keywords[Environment]', 'Environmental Keywords[Certificate]', 'Environmental Keywords[Green Practices]',
                 'Environmental Keywords[Sustainable Transportation]', 'Social Keywords', 'Cultural Keywords', 'Economic Keywords',
                 'Policy Keywords'])


print("Processing...top enviornmental keywords subcatergories")
top_10_enviornment = get_top_10_keywords(df2,'Environmental Keywords[Environment]')
top_10_enviornment_list = [str(item) for item in  top_10_enviornment]  # convert tuple to list of strings
top_10_certificate = get_top_10_keywords(df2,'Environmental Keywords[Certificate]')
top_10_certficate_list = [str(item) for item in  top_10_certificate]  # convert tuple to list of strings
top_10_green_practices = get_top_10_keywords(df2,'Environmental Keywords[Green Practices]')
top_10_green_practices_list = [str(item) for item in top_10_green_practices]  # convert tuple to list of strings
top_10_sustainable_transportation = get_top_10_keywords(df2,'Environmental Keywords[Sustainable Transportation]')
top_10_sustainable_transportation_list  = [str(item) for item in top_10_sustainable_transportation]  # convert tuple to list of strings

print("Processing...top social keywords")
top_10_social = get_top_10_keywords(df2,'Social Keywords')
top_10_social_list = [str(item) for item in top_10_social]  # convert tuple to list of strings
print("Processing...top cultural keywords")
top_10_cultural = get_top_10_keywords(df2,'Cultural Keywords')
top_10_cultural_list = [str(item) for item in top_10_cultural]  # convert tuple to list of strings
print("Processing...top economic keywords") 
top_10_economic = get_top_10_keywords(df2,'Economic Keywords')
top_10_economic_list = [str(item) for item in top_10_economic]  # convert tuple to list of strings
print("Processing...top policy keywords") 
top_10_policy = get_top_10_keywords(df2,'Policy Keywords')
top_10_policy_list = [str(item) for item in top_10_policy]  # convert tuple to list of strings

print('Writing to dic')
# Append top 10 keywords from each category to dictionary
# itertools.zip_longest() function to iterate over all lists, even if they have different lengths, value of "" to fill any missing values.
for enviornment_keywords, certificate_keywords, green_practices_keywords, sustainable_transportation_keywords, social_keywords, cultural_keywords, economic_keywords, policy_keywords in itertools.zip_longest(
           top_10_enviornment_list, top_10_certficate_list,  top_10_green_practices_list, top_10_sustainable_transportation_list, top_10_social_list, top_10_cultural_list, top_10_economic_list, top_10_policy_list,
            fillvalue=""):
        research_ques2['Top 10 Environmental Keywords[Environment]'].append(enviornment_keywords)
        research_ques2['Top 10 Environmental Keywords[Certificate]'].append(certificate_keywords)
        research_ques2['Top 10 Environmental Keywords[Green Practices]'].append(green_practices_keywords)
        research_ques2['Top 10 Environmental Keywords[Sustainable Transportation]'].append(sustainable_transportation_keywords)        
        research_ques2['Top 10 Social Keywords'].append(social_keywords)
        research_ques2['Top 10 Cultural Keywords'].append(cultural_keywords)
        research_ques2['Top 10 Economic Keywords'].append(economic_keywords)
        research_ques2['Top 10 Policy Keywords'].append(policy_keywords)

update_excel(research_ques2, worksheet, skip_columns_after=['Top 10 Policy Keywords'])
