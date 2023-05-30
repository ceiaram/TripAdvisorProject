import pymongo
import dns # required for connecting with SRV
import pandas as pd
import spacy
import re
import csv
import openpyxl
from collections import Counter
# Import the Matcher
from spacy.matcher import PhraseMatcher
from spacy.matcher import Matcher
from openpyxl.utils.exceptions import IllegalCharacterError
from lxml import etree

# https://course.spacy.io/en/chapter1

# Connect to database, 24hr timeout
client = pymongo.MongoClient("mongodb+srv://ceiaram:Xn704K3WQ0So6s09@tripadvisordb.qw0eojr.mongodb.net/?retryWrites=true&w=majority", socketTimeoutMS=86400000)
db = client.test

# Get database named "OrangeCountyHotels"
db = client["OrangeCountyHotels"]

# Get collection
collection_reviews = db['Reviews']
collection_hotels = db['HotelData']

# Create a new workbook
workbook = openpyxl.Workbook()
# Select the active worksheet
worksheet = workbook.active

# All the scores found in each category(reviews)
all_scores = []
# About section score
about_section_scores = []
# Total hotel scores from each category 
total_hotel_scores = []

# Analyze syntax
class Analysis:
    def getNounPhrases(doc):
        noun_phrases = [chunk.text for chunk in doc.noun_chunks]
        return noun_phrases
    def getVerbs(doc):
        verbs = [token.lemma_ for token in doc if token.pos_ == "VERB"]
        return verbs
    def getCommonWords(doc, numOfMostCommon):
        words = [ token.text for token in doc if not token.is_stop and not token.is_punct]
        common_words = Counter(words).most_common(numOfMostCommon)
        return common_words
    def getLabels(doc):
        labels = {}
        # PERSON - People, including fictional.
        # NORP - Nationalities or religious or political groups.
        # FAC - Buildings, airports, highways, bridges, etc.
        # ORG - Companies, agencies, institutions, etc.
        # GPE - Countries, cities, states.
        # Find named entities, phrases and concepts
        for entity in doc.ents:
            labels[entity.text] = entity.label_
            # print(entity.text, entity.label_)
        return labels
    
    # This code fixes replaces illegal characters with valid ones
    def update_excel(data_dict, worksheet, skip_column_after=None):
        # Clear the existing data in the worksheet
        worksheet.delete_rows(1, worksheet.max_row)

        # Write the headers to the first row
        headers = list(data_dict.keys())
        for i, header in enumerate(headers):
            col_index = i + 1
            if skip_column_after is not None and i >= headers.index(skip_column_after):
                col_index += 2 #Create 2 blank columns at specific key 
            worksheet.cell(row=1, column=col_index).value = header

        # Pad the shorter lists with empty values
        max_length = max([len(data_dict[key]) for key in headers])
        for key, value in data_dict.items():
            if len(value) < max_length:
                data_dict[key] += [''] * (max_length - len(value))

        # Write the data to the remaining rows
        for i in range(max_length):
            for j, header in enumerate(headers):
                col_index = j + 1
                if skip_column_after is not None and j >= headers.index(skip_column_after):
                    col_index += 2 #Create 2 blank columns at specific key 
                value = data_dict[header][i]
                try:
                    worksheet.cell(row=i+2, column=col_index).value = value
                except IllegalCharacterError as e:
                    print(f"\nIllegal character found in {header} column, row {i+2}: {e.args[0]}")
                    # Replace all illegal characters with underscores
                    value = ''.join(['_' if ord(char) < 32 or ord(char) > 126 else char for char in e.args[0]])
                    worksheet.cell(row=i+2, column=col_index).value = value

        # Check if there are any invalid XML elements in the workbook
        with open("my_workbook.xlsx", "rb") as f:
            try:
                etree.parse(f)
            except etree.XMLSyntaxError as e:
                print(f"\nInvalid XML element found in workbook: {e.args[0]}")

        # Save the workbook
        workbook.save("my_workbook.xlsx")


    def findKeywords(text, keywords, str_list_name):
       # create a regular expression pattern to match the whole word that contains one of the keywords(ex: Disney should find Disneyland and Disney-land and landdisney)
       pattern2 = re.compile(r'\b(\w*{}[\w]*)\b'.format('|'.join(re.escape(keyword) for keyword in keywords)), re.IGNORECASE)

       # create a regular expression pattern matches any word that starts with one of the keywords and is followed by zero or more word characters
       # main difference with previous pattern is this pattern does not capture the matched text, re.IGNORECASE is for disregard case sensitivity 
       pattern3 = re.compile(r'\b(?:{})\w*\b'.format('|'.join(map(re.escape, keywords))), re.IGNORECASE)
        
       word_variations = []
       #store the two-worded keywords and their respective values
       two_worded_keywords = {}

       # Includes two-worded keywords in the dictionary and adds their value to the score if both words are found in the text
       for keyword in keywords:
        words = keyword.split()
        if len(words) > 1:
            first_word = words[0]
            second_word = words[1]
            word_variations.append(r'\b{}\w*\b.*?\b{}\w*\b|\b{}\w*\b \w*\b'.format(re.escape(keyword), re.escape(second_word), re.escape(first_word)))
            two_worded_keywords[keyword] = keywords[keyword]

       # Pattern matches either a single word from the dictionary keys, or a two-word phrase that has been split into its constituent words and combined with the \w*
       pattern = re.compile(r'\b(?:{})\w*\b|\b(?:{})\b'.format('|'.join(map(re.escape, keywords.keys())), '|'.join(word_variations)), re.IGNORECASE)

       # Scores for each category 
       score = 0

       # search for a keyword in the text from pattern
       match = pattern.findall(text)
    #    Don't need pattern 2 or 3
    #    match2 = pattern2.findall(text)
    #    match3= pattern3.findall(text)

       matchesFound = []  
       matchesFound.extend(match)
    #    matchesFound.extend(match2)
    #    matchesFound.extend(match3)

       # remove duplicates from the list of keyword matches
       matchesFound = list(set(matchesFound))

       printed_matches = set()

       # Loop over all matches found and add correct values based on keywords
       for match in matchesFound:
            # Check if the match is a one-worded keyword and add its value to the score
            for keyword, value in keywords.items():
                if keyword in match.lower() and match not in printed_matches:
                    score += value
                    print(f"Matched keyword '{keyword}' with value {value}: '{match}'")
                    printed_matches.add(match)
            # Check if the match is a two-worded keyword and add its value to the score 
            for keyword, value in two_worded_keywords.items():
                if keyword in match.lower() and match not in printed_matches:
                    score += value
                    print(f"Matched keyword '{keyword}' with value {value}: '{match}'")
                    printed_matches.add(match)

       # Append scores to correct list to be added 
       if str_list_name == "review":
        all_scores.append(score)
       elif str_list_name == "about":
        about_section_scores.append(score)
           
       if printed_matches:
        print("Keywords found: ", printed_matches)
        print("Score: ", score)
       else:
        print('No keywords found.')
        printed_matches.add('none')
        
       return printed_matches


if __name__ == "__main__":
    
    # Load English tokenizer, tagger, parser and NER
    nlp = spacy.load("en_core_web_sm")

    # Initialize the phrase matcher with the shared vocabulary
    phrase_matcher = PhraseMatcher(nlp.vocab)
    matcher = Matcher(nlp.vocab)

    # Create a dictionary with initially with no data
    dict = { 
        'Hotel Name':[],  'Hotel About Section':[],  'Hotel About Section Keywords':[], 'Hotel About Section Sustainabilty Score' :[],
        'Hotel Rating':[], 'Hotel Class' :[], 'Hotel Price Range' :[], 'Total Number of Reviews':[], 
        'Hotel Value Score':[], 'Reviews' : [], 'Manager Response':[], 'Enviornmental Keywords' : [], 'Environmental Keywords[Environment]' : [] , 'Environmental Keywords[Certificate]' : [],
        'Environmental Keywords[Green Practices]' : [], 'Environmental Keywords[Sustainable Transportation]' : [],
        'Social Keywords': [], 'Cultural Keywords' : [],'Economic Keywords' : [], 'Policy Keywords' : [], 'Top 5 Common Words':[], 'Labels':[], 'Customer Rating':[],
        'Date of Stay':[], 'Helpful Vote':[], 'Sustainabilty Score' : [], 'Total Sustainable Keywords Found' : [], 'Hotel Sustainabilty Average':[]
    }
 
 
    # Weights are all 1 
    # HAVE ALL KEYWORDS LOWERCASE TO FIND SCORES
    environmental_keywords = {
    'automatic light': 1,   # TWO-WORDED KEYWORDS MUST COME BEFORE ONE-WORDED KEY WORDS THAT ARE FOUND IN THE TWO-WORDED KEYWORDS OTHERWIS IT WON'T PICK UP THE TWO-WORDED KEYWORDS/CCORRECT SCORES 
    'clean air': 1,  
    'clean environment': 1, 
    'energy sav': 1, ##
    'environmental certificat': 1, 
    'environmental concern': 1, 
    'environmental manage': 1, 
    'environmentally friendly': 1, 
    'ev charg': 1, 
    'ev park': 1, 
    'electronic car': 1, 
    'electric car': 1, 
    'iso 14001': 1, 
    'towel reuse': 1, 
    'linen reuse' : 1, 
    'led light': 1, 
    'leed certificat': 1, 
    'locally grown': 1, 
    'motion sensor': 1, 
    'green space': 1, 
    'green globe': 1, 
    'green hotel': 1, 
    'green issue': 1, 
    'green lodge': 1, 
    'green practice': 1, 
    'green certificat': 1, 
    'recharge station': 1, 
    'sustainable material': 1, 
    'sustainable land': 1, 
    'sustainable transportation': 1, 
    'waste conservat': 1, 
    'water sav': 1, 
    'vehicle charg': 1, 
    # 'green': 1,
    'conservat': 1, 
    'ecological ': 1, 
    'eco-friendly': 1, 
    # 'environment': 1, 
    # 'reuse': 1,
    'recycle': 1, 
    'renewable': 1, 
    'solar': 1, 
    'organic': 1, 
    'biodegrad': 1, 
    'biodiversit': 1, 
    'restorat': 1, 
    'hybrid': 1, 
    'natur': 1, 
    'efficient': 1, 
    'vegetarian': 1, 
    'vegan': 1, 
    'healthy': 1, 
    'sustainability': 1, 
    # 'leed': 1,
    'non-fuel': 1, 
    'non fuel' : 1, 
    'reduce': 1 
    }

    environmental_keywords_environment = {
        'clean air': 1,  
        'clean environment': 1, 
        'green space': 1, 
        'green globe': 1, 
        'green hotel': 1, 
        'green issue': 1, 
        'green lodge': 1, 
        'green practice': 1, 
        'sustainable land': 1,
        'ecological ': 1, 
        'sustainability': 1,
    }

    environmental_keywords_certificate = {
        'environmental certificat': 1,
        'leed certificat': 1,
        'green certificat': 1,
        'iso 14001': 1,
    }

    environmental_keywords_practices = {
        'automatic light': 1,
        'environmental concern': 1, 
        'environmentally friendly': 1,
        'environmental manage': 1,
        'towel reuse': 1, 
        'led light': 1,
        'linen reuse' : 1,
        'locally grown': 1,
        'energy sav': 1,
        'motion sensor': 1,
        'sustainable material': 1,
        'waste conservat': 1,
        'water sav': 1,
        'eco-friendly': 1,
        'recycle': 1,
        'renewable': 1,
        'conservat': 1,
        'solar': 1,
        'organic': 1,
        'biodegrad': 1,
        'biodiversit': 1,
        'restorat': 1,
        'natur': 1,
        'efficient': 1,
        'vegetarian': 1,
        'vegan': 1,
        'healthy': 1,
        'reduce': 1
    }

    environmental_keywords_sustainable_transportation = {
        'ev charg': 1,
        'ev park': 1,
        'electronic car': 1,
        'electric car': 1,
        'sustainable transportation': 1,
        'recharge station': 1,
        'non fuel' : 1,
        'vehicle charg': 1,
        'non-fuel': 1,
        'hybrid': 1,
    }


    social_keywords = {
    'corporate responsibilit': 1,
    'community support': 1,
    'ethical practice': 1,
    'local employee': 1,
    'local purchase': 1,
    'local communit': 1,
    'responsible consumer behavior': 1,
    'responsible customer behavior': 1,
    'social responsibilit': 1,
    'sustainable communit': 1,
    # 'communit': 1,
    # 'responsibl': 1,
    # 'involvement': 1,
    }

    cultural_keywords = {
    'cultural sustainabilit': 1,
    'safe work': 1,
    'work safe' : 1,
    'heritage': 1,
    'inclusi': 1,
    'traditional': 1,
    'diversit': 1,
    'equit': 1,
    # 'fairly': 1,
    'fairness' : 1,
    'dei': 1,
    }

    economic_keywords = {
    'economical' : 1,
    'market demand': 1,
    'competitive': 1,
    'advantage': 1,
    'innovat': 1,
    'profit': 1,
    'invest': 1,
    'value': 1,
    # 'renovat': 1
    }

    policy_keywords = {
    'corporate social responsibilit': 1,
    'environmental regulation': 1,
    'labor regulat': 1,
    'organizational identit': 1,
    'social responsibilit': 1,
    'social justice': 1,
    'tax polic': 1,
    'zoning polic': 1,
    'stakeholder': 1,
    'regulation': 1,
    'policy': 1, 
    'policies': 1,
    'transparen': 1   
    }

    # All reviews from 587 hotels (500-587 are good) (0-10)
    # Error at index: 308, 400??403  DoubleTree by Hilton Buena Park
    for x in range(169, 180):
        # Find Hotels and Reviews at specific index (0 is the first hotel and first hotel's reviews...
        # ..1 is the second hotel and the second hotel's reviews and so on...)
        my_query = {'Index' : x}

        # Check if reviews query exists, if so perform data analysis 
        if(collection_reviews.find_one(my_query)):
            # Get review documents and hotel documents
            review_doc = collection_reviews.find(my_query)
            hotel_doc = collection_hotels.find(my_query)

            # Print Hotel with query 
            for doc in hotel_doc:
                # Obtain the hotel fields for the each hotel 
                hotel_name = doc['Hotel Name']
                hotel_rating = doc['Hotel Rating']
                hotel_class = doc['Hotel Class']
                price_range = doc['Hotel Price Range']
                # total_num_reviews = doc['Total Number of Reviews']
                value_score = doc['Hotel Value Score']
                if doc['Hotel About Section']:
                    about_section = doc['Hotel About Section']
                    # if the about section is a string, access its elements
                    if isinstance(about_section, str):
                        #Find keywords in hotel about section from each category of dictionary 
                        enviornmentalResultsHotelDes = Analysis.findKeywords(about_section, environmental_keywords, "about")
                        socialResultsHotelDes = Analysis.findKeywords(about_section, social_keywords, "about")
                        culturalResultsHotelDes = Analysis.findKeywords(about_section, cultural_keywords, "about")
                        economicResultsHotelDes = Analysis.findKeywords(about_section, economic_keywords, "about")
                        policyResultsHotelDes = Analysis.findKeywords(about_section, policy_keywords, "about")

                        # Get the sum of all the scores found in each category for the about section 
                        sum_scores_about = 0
                        print("ABOUT SECTION SCORES", about_section_scores)
                        for scores in about_section_scores:
                            sum_scores_about+= scores
                else:
                    about_section = "none"
                # Clear list for next hotel 
                about_section_scores.clear()
        
            # Obtain reviews with query 
            for doc in review_doc:
                # Obtain the review fields for the each review  
                summary = doc['Summary']
                review_num = doc['Review Number']
                customer_rating = doc['Overall Rating']
                date_of_stay = doc['Date of Stay']
                helpful_vote = doc['Helpful Vote']
                manager_response = doc['Manager Response']

                # Perform Data Analysis on each review of hotel x
                text = (summary)
                doc_nlp = nlp(text)

                print("HOTEL NAME: ", hotel_name, " at index", x)
                print(review_num)
                print(summary)
    
                # Data analysis using the spacy library 
                commonWords = Analysis.getCommonWords(doc_nlp, 5)
                commonWordsList = [str(item) for item in commonWords]  # convert tuple to list of strings
                labels = Analysis.getLabels(doc_nlp)
                print(labels)
                labelsList =  [f"{k}: {v}" for k, v in labels.items()] # convert dictionary to list of strings using f-string

                # Find key words in reviews from each category of dictionary
                enviornmentalResults = Analysis.findKeywords(text, environmental_keywords, "review")
                # Sub-catergories of enviornmental dict, so keywords don't need to be added 
                enviornmentalResultsEnvironment = Analysis.findKeywords(text, environmental_keywords_environment, None)
                enviornmentalResultsCertificate = Analysis.findKeywords(text, environmental_keywords_certificate, None)
                enviornmentalResultsGreenPractices = Analysis.findKeywords(text, environmental_keywords_practices, None)
                enviornmentalResultsSustainableTransportation = Analysis.findKeywords(text, environmental_keywords_sustainable_transportation, None)
  


                socialResults = Analysis.findKeywords(text, social_keywords, "review")
                culturalResults = Analysis.findKeywords(text, cultural_keywords, "review")
                economicResults = Analysis.findKeywords(text, economic_keywords, "review")
                policyResults = Analysis.findKeywords(text, policy_keywords, "review")
      
                # Get the sum of all the scores found in each category for reviews 
                sum_scores = 0
                print("BEFORE LOOP", all_scores)
                for scores in all_scores:
                    sum_scores+= scores
                print("AFTER LOOP", sum_scores)
             
                # Add results of review to dictionary 
                dict['Reviews'].append(summary)
                dict['Manager Response'].append(manager_response)
                dict['Enviornmental Keywords'].append(", ".join(enviornmentalResults))
                dict['Environmental Keywords[Environment]'].append(", ".join(enviornmentalResultsEnvironment))
                dict['Environmental Keywords[Certificate]'].append(", ".join(enviornmentalResultsCertificate))
                dict['Environmental Keywords[Green Practices]'].append(", ".join(enviornmentalResultsGreenPractices))
                dict['Environmental Keywords[Sustainable Transportation]'].append(", ".join(enviornmentalResultsSustainableTransportation))

                dict['Social Keywords'].append(", ".join(socialResults))
                dict['Cultural Keywords'].append(", ".join(culturalResults))
                dict['Economic Keywords'].append(", ".join(economicResults))
                dict['Policy Keywords'].append(", ".join(policyResults))
                dict['Top 5 Common Words'].append(", ".join(commonWordsList))
                dict['Labels'].append(", ".join(labelsList))
                dict['Customer Rating'].append(customer_rating)
                dict['Date of Stay'].append(date_of_stay)
                dict['Helpful Vote'].append(helpful_vote)

                # Add hotel data to dictionary, inside review for loop to fill all cells
                dict['Hotel Name'].append(hotel_name)
                dict['Hotel Rating'].append(hotel_rating)
                dict['Hotel Class'].append(hotel_class)
                dict['Hotel Price Range'].append(price_range)
                # dict['Total Number of Reviews'].append(total_num_reviews)
                dict['Hotel Value Score'].append(value_score)
                # if the about section is a string, access its elements
                if isinstance(about_section, str):
                    dict['Hotel About Section'].append(about_section)
                    dict['Hotel About Section Keywords'].append(", ".join(list(enviornmentalResultsHotelDes) + list(socialResultsHotelDes) 
                                                                        + list(culturalResultsHotelDes) + list(economicResultsHotelDes) + list(policyResultsHotelDes)))
                else:
                     # otherwise, handle the NaN value or other non-string values
                    dict['Hotel About Section'].append("none")
                    dict['Hotel About Section Keywords'].append("none")

                dict['Hotel About Section Sustainabilty Score'].append(sum_scores_about)

                print("BEFORE INSERTION",sum_scores)
                # Add sustainabilty score to dictionary 
                dict['Sustainabilty Score'].append(sum_scores)


                # Append all the scores from each category to total hotel scores 
                total_hotel_scores.append(sum_scores)
                # Clear list for next review analysis
                all_scores.clear()

                
            # Calculate the average if the list is not empty
            if len(total_hotel_scores) > 0:
                # Calculate the sum of total hotel scores 
                sum_total_hotel_scores = sum(total_hotel_scores)
                # Calculate the average
                average = sum_total_hotel_scores / len(total_hotel_scores)
                rounded_average = round(average, 4)
                print(rounded_average)
            else:
                # set rounded avg in the empty list case
                rounded_average = 0.0000

            # Total number of reviews found for hotel 
            total_num_reviews = len(total_hotel_scores)
            for i in total_hotel_scores:
                   dict['Total Number of Reviews'].append(total_num_reviews)
                   print("HEEEEELEL")


            # Clear list 
            total_hotel_scores.clear()

            # Append the rounded average and total to the dictionary
            dict['Total Sustainable Keywords Found'].append(sum_total_hotel_scores)
            dict['Hotel Sustainabilty Average'].append(rounded_average)
            # Update excel file with data     
            Analysis.update_excel(dict, worksheet, skip_column_after="Reviews")
        else:
            print("Index ", x , "does not exist")
            continue

        # Print the number of docuements with specific index
        # print(collection_reviews.count_documents(my_query))
        # print(collection_hotels.count_documents(my_query))


    

